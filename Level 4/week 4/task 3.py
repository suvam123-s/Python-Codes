import tkinter as tk 
from tkinter import ttk
import openpyxl
from openpyxl import Workbook


root= tk.Tk()
root.title("Data Store System")
root.geometry("400x400")

FILE_NAME="class_work2.xlsx"

def initializeExcel():
    try:
        workbook= openpyxl.load_workbook(FILE_NAME)
        worksheet=workbook.active
        print("file found logic")
    except:
        workbook= Workbook()
        worksheet=workbook.active
        worksheet.append(["Name","age"])
        workbook.save(FILE_NAME)
        print("no file found")

ui_frame=tk.Frame(root)
ui_frame.grid(row=1,column=0)


tk.Label(ui_frame,text="Name").grid(row=0,column=0)
entry_name = tk.Entry(ui_frame)
entry_name.grid(row=0,column=1)

tk.Label(ui_frame,text="Age").grid(row=1,column=0)
entry_age= tk.Entry(ui_frame)
entry_age.grid(row=1,column=1)

# def tableInsert(*args):
#     table.insert("",tk.END,values=args)

def tableInsert(**kwargs):
    print(kwargs['name'])
    print(kwargs['age'])
    table.insert("",tk.END,values=(kwargs['name'],kwargs['age']))

def save_data():
    name=entry_name.get()
    age=entry_age.get()
    tableInsert(name=name,age=age)

    workbook=openpyxl.load_workbook(FILE_NAME)
    worksheet=workbook.active
    worksheet.append((name,age))
    workbook.save(FILE_NAME)
    # print(workbook)

    # print(name,age)



tk.Button(ui_frame,text="Save",command=save_data).grid(row=2,column=1)

table=ttk.Treeview(ui_frame, columns=("Name","Age"), show="headings")
table.heading("Name", text="Name")
table.heading("Age",text="Age")
table.grid(row=3,column=0,columnspan=2)

initializeExcel()

root.mainloop()